
root_c_paths = {
    'Windows':  'C:\\',
    'Linux':    '/mnt/c'
}
project_path_subdirs = ['Users', 'miken', 'Rice', 'Research', 'gut_mb_autoencoder']


import os, sys, datetime, argparse, logging, platform
# import phylogeny_utilities.utilities as phy
import openpyxl
from settings_configs import *

project_path = os.path.join(root_c_paths[platform.system()], *project_path_subdirs)
scripts_path = os.path.join(project_path, 'scripts')

excel_master_filename = 'Master_Project_Dataset_Inventory_20231223.xlsx'
excel_master_path = os.path.join(project_path, excel_master_filename)
excel_bioproject_tab = 'Total_Storage_Estimate'




parser = argparse.ArgumentParser()
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

def parse_command_args():
    global parser
    # parser.add_argument('-v','--verbose',dest='verbose', action='store_true', help='prints more detailed output to console.')
    # parser.add_argument('--debug', dest='debug', action='store_true')

    subparsers = parser.add_subparsers(help='specifies the action to take.')
    update_bp_location_reference = subparsers.add_parser('update_bp_location_reference',
                                         help='Updates the file containing the list of bioprojects and their storage '
                                              'folders on the server. By default this file is called `target_bioproject_path_list.tsv` '
                                              'and is locaed in the `scripts` subfolder.')
    update_bp_location_reference.add_argument('-o','--output', dest='output_file', type=str,
                              default=os.path.join(scripts_path, bioproject_storage_path_list_filename),
                              help='path to the output file to be written to as a tsv.')
    update_bp_location_reference.add_argument('-d', '--dump_tables', dest='dump_tables', default=False, action='store_true',
                                              help='if given, the key reference tables will be written to text'
                                                   ' files in the scripts subfolder.')
    update_bp_location_reference.add_argument('--also_dump_sra_list', dest='dump_sralist', default=False, action='store_true',
                                              help='if given, will also dump the big table to a text file by the default ' \
                                                   'name.')
    update_bp_location_reference.set_defaults(func=update_bioproject_location_reference, command='update_references')

    make_globus_batch_file = subparsers.add_parser('make_globus_commands_file')
    make_globus_batch_file.add_argument('-o','--output', dest='output_file', type=str, required=True)
    make_globus_batch_file.add_argument('-bp', '--bioprojects', dest='bioprojects', type=str, required=True)
    make_globus_batch_file.set_defaults(func=run_make_globus_batch_file, command='make_globus_batch')


    dump_excel_runsdata = subparsers.add_parser('dump_excel_runsdata_to_text')
    dump_excel_runsdata.add_argument('-o', '--output', dest='output_file', type=str,
                                        default=sra_runslist_default_dumpfile)
    dump_excel_runsdata.set_defaults(func=dump_excel_runlist_to_text, command='dump_sra_list')

    args = parser.parse_args()
    return args

def dump_excel_runlist_to_text(cmd_args, workbook=None):
    '''
    Function to dump the big table of SRA runs from excel into a single tabular format.

    Crucially this function doesn't use any fancy save-as command. It just reads the cells row-by-row and
    writes them over to text.

    :param cmd_args:    command-line arguments
    :param workbook:    copy of workbook opbject representing this workbok.
    :return:
    '''
    logger.info(f'Opening excel workbook: {excel_master_filename}')
    if workbook is None:
        wb = openpyxl.load_workbook(filename=excel_master_path, data_only=True)
    else:
        wb = workbook
    data_sheetname = 'SraRunList_data'


    out_file = cmd_args.output_file
    if cmd_args.command == 'update_references':
        out_file = sra_runslist_default_dumpfile
    out_path = os.path.join(project_path, out_file)

    with open(out_path,'w') as out_f:
        for r in wb[data_sheetname].rows:
            cct = out_f.write('\t'.join([str(i.value) for i in r]) + '\n')

    wb.close()
    pass

def get_storage_loc_name_path_lookup(wb, write_to_file=False):
    '''
    Gets the lookup table from Excell that maps the short-names to the folders on
    each hard drive that we happen to be storing there.
    '''
    t, c = next(wb.defined_names['ref_storage_locations'].destinations)
    ref_locs = {i[0].value: i[1].value for i in wb[t][c] if i[0].value is not None}
    if write_to_file:
        with open(os.path.join(scripts_path, 'ref_storage_locations.txt'),'w') as rsl:
            cct = rsl.write('loc_name\tloc_path\n')
            for i in ref_locs:
                cct = rsl.write(f'{i}\t{ref_locs[i]}\n')
    return ref_locs

def get_bioproject_storage_loc_lookup(wb, write_to_file=False):
    '''
    Gets the lookup table from Excel that maps bioproject to which hard-drive locations
    the files are kept in (expressed as short name, of course).
    '''
    myws = wb[excel_bioproject_tab]
    # Get list of bioprojects:
    bp_title, bp_coord = next(wb.defined_names['ref_bioproject_list'].destinations)
    bioprojects = [ i[0].value for i in myws[bp_coord] ]
    bp_row = myws[bp_coord][0][0].row; bp_col = myws[bp_coord][0][0].column;
    # Get list of table headers:
    bioproject_tbl_hdrs = []
    for mycol in range(bp_col, myws.max_column + 1):
        bioproject_tbl_hdrs.append(myws.cell(bp_row-1, mycol).value)
    #
    storage_loc_col = bioproject_tbl_hdrs.index('storage_loc') + bp_col
    storage_locs = [myws.cell(bp_row + i, storage_loc_col).value for i in range(len(bioprojects))]
    #
    if write_to_file:
        other_cols = ['n_runs','storage_loc','ena_total_file_size','ena_n_files']
        col_inds = [bp_col,] + [bioproject_tbl_hdrs.index(i) + bp_col for i in other_cols]
        data = [[myws.cell(bp_row + i, col).value for col in col_inds]
                for i in range(len(bioprojects))]
        outpath = os.path.join(scripts_path, 'ref_bioproject_data.tsv')
        with open(outpath,'w') as out_f:
            cct = out_f.write('bioproject\t' + '\t'.join(other_cols) + '\n')
            for d in data:
                cct = out_f.write('\t'.join([str(val) for val in d])+'\n')

    # dict from bioproject to storage loc alias: { <bioproject>: <storage_loc_name> }
    return dict(zip(bioprojects, storage_locs))

def update_bioproject_location_reference(cmd_args):
    '''Goes through the excel file and writes a tab-separated list of bioprojects along with the root folder
    the raw data is stored in on the server. Works by getting from excel the shortname-to-path lookup, then the
    bioproject-to-shortname lookup, then writing the output as a tsv.'''
    logger.info(f'Opening excel workbook: {excel_master_filename}')
    wb = openpyxl.load_workbook(filename=excel_master_path, data_only=True)

    # First two columns of range 'ref_storage_locations' on tab 'ref':
    logger.info(f'getting storage-shortname-to-path lookup.')
    storage_path_name_lkp = get_storage_loc_name_path_lookup(wb, write_to_file=cmd_args.dump_tables)
    # Column 1 and column with header 'storage_loc' from main table on tab 'Total_Storage_Estimate' (stored in a
    #   variable earlier.
    logger.info(f'getting bioproject to storage-shortname lookup.')
    BP_to_storage_loc = get_bioproject_storage_loc_lookup(wb, write_to_file=cmd_args.dump_tables)

    nonempty_bioprojects = [k for k in BP_to_storage_loc if BP_to_storage_loc[k] is not None]
    logger.info(f'Found {len(nonempty_bioprojects)} projects with a specified storage location. Writing to file.')
    target_paths_file = cmd_args.output_file

    with open(target_paths_file, 'w') as tpf:
        for bp in nonempty_bioprojects:
            spath = storage_path_name_lkp[BP_to_storage_loc[bp]]
            cct = tpf.write(f'{bp}\t{spath}\n')

    if cmd_args.dump_sralist:
        dump_excel_runlist_to_text(cmd_args, wb)

    wb.close()

def parse_sralist():
    '''Reads the big SRA Runlist file (specifically that has just been exported
    from the big excel file. Nothing fancy here, just a big specific csv-parser.'''
    import csv
    sra_list_filename = 'Master_Project_Dataset_Runlist.txt'
    sra_path = os.path.join(project_path, sra_list_filename)
    with open(sra_path,newline='') as sra:
        sracsv=csv.reader(sra,delimiter='\t')
        srarows = [i for i in sracsv]
    headers = srarows[0]
    bp_column = [i.lower() for i in headers].index('bioproject')
    bps = list(set([i[bp_column] for i in srarows[1:]]))
    res = {bp: [] for bp in bps}
    for d in srarows[1:]:
        res[d[bp_column]].append(d)
    return headers, res

def tsv_to_tuples(filepath):
    with open(filepath,'r') as f:
        lns = [i.strip().split('\t') for i in f.readlines()]
    return lns

def parse_ena_metadata():
    ena_metadata_folder = os.path.join(project_path, 'ena_bioproject_metadata')
    ena_files = os.listdir(ena_metadata_folder)
    bp_list = [i.replace('.txt','') for i in ena_files]
    res = {}
    for f in ena_files:
        ftup = tsv_to_tuples(os.path.join(ena_metadata_folder, f))
        res[f.replace('.txt', '')] = {i[0]: [tuple(j.split(';')) for j in i[1:]] for i in ftup[1:]}
    return res

def run_make_globus_batch_file(cmd_args):
    sra_hdrs, sra_runs = parse_sralist()
    ena = parse_ena_metadata()

    bp_list = cmd_args.bioprojects.split(',')
    print(bp_list)
    assert set(bp_list)==set(bp_list) & set(ena.keys()), f'BioProjects list includes some not in our list: {str(bp_list)}'

    trim_ftp_path = lambda x: x.split('fastq/')[1]
    globus_cmd = open(os.path.join(project_path,'download_scripts', cmd_args.output_file),'w')

    for bp in bp_list:
        ks_ena = ena[bp].keys()
        ks_sra = [i[0] for i in sra_runs[bp]]
        ks = list(set(ks_sra) & set(ks_ena))
        for runid in ks:
            files = [trim_ftp_path(f) for f in ena[bp][runid][0]]
            for f in files:
                cct = globus_cmd.write(f'{f} {bp}/{os.path.split(f)[1]}\n')

    globus_cmd.close()






    pass

if __name__ == '__main__':
    cmd_args = parse_command_args()
    cmd_args.func(cmd_args)

